# -*- coding: utf-8 -*-

# 写一个新文件
# from openpyxl.workbook import Workbook

# data = [
#     ['气瓶编号', '分类', '站点编号', '性质', '状态', '添加时间', '压力', 'primary浓度'],
#     ['CBO9119', '', '52134', '', '空瓶', '2018-07-05', '45.6', '1423.62'],
#     ['CBO9120', '', '52135', '', '报废瓶', '2018-07-05', '42.9', '1523.62'],
#     ['CBO9121', '', '52136', '', '报废瓶', '2018-07-05', '44.9', '1473.62'],
#     ['CBO9122', '', '52137', '', '报废瓶', '2018-07-05', '44.5', '1363.62'],
#     ['CBO9123', '', '52138', '', '报废瓶', '2018-07-05', '43.5', '1363.78']
# ]

# 在内存创建一个工作簿对象
# wb = Workbook()

# 设定第一个sheet页
# ws = wb.active
# ws.title = u'瓶子信息'

# 向第一个sheet页写数据
# i = 1
# for line in data:
#     for col in range(1, len(line)+1):
#         ws.cell(column=col, row=i, value=line[col-1])
#     i += 1

# 工作薄保存到磁盘
# wb.save('瓶子信息.xlsx')

# --------------------------------------------------

# 写一个已经存在的文件
from openpyxl import load_workbook

data = [
    # ['气瓶编号', '分类', '站点编号', '性质', '状态', '添加时间', '压力', 'primary浓度'],
    ['CBO9119', '', '52134', '', '空瓶', '2018-07-05', '45.6', '1423.62'],
    ['CBO9120', '', '52135', '', '报废瓶', '2018-07-05', '42.9', '1523.62'],
    ['CBO9121', '', '52136', '', '报废瓶', '2018-07-05', '44.9', '1473.62'],
    ['CBO9122', '', '52137', '', '报废瓶', '2018-07-05', '44.5', '1363.62'],
    ['CBO9123', '', '52138', '', '报废瓶', '2018-07-05', '43.5', '1363.78']
]

path = r"D:\pywork\resource\添加新瓶模板.xlsx";
wb = load_workbook(path)

ws = wb[wb.sheetnames[0]]
ws.title = "瓶子"

# 让出表头位置
i = 2
for line in data:
    for col in range(1, len(line)+1):
        # 两种写法都行
        # ws.cell(row=i, column=col, value=line[col-1])
        ws.cell(row=i, column=col).value = line[col-1]
    i += 1

wb.save(path)

